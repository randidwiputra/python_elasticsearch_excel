from elasticsearch import Elasticsearch
import dateutil.parser

import openpyxl
from openpyxl import Workbook

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


#wb = load_workbook('test_Excel2.xlsx')
wb =  Workbook()
sheet = wb.active



sheet['A1'] = 'asalKota'
sheet['B1'] = 'awbNumber'
sheet['C1'] = 'lastValidTrackingDateTime'
sheet['D1'] = 'lastValidTrackingSiteName'
sheet['E1'] = 'lastValidTrackingType'
sheet['F1'] = 'layanan'
sheet['G1'] = 'manifestDateTime'
sheet['H1'] = 'manifestTrackingSiteCity'
sheet['I1'] = 'manifestTrackingSiteName'
sheet['J1'] = 'noKonfirmasi'
sheet['K1'] = 'perwakilan'
sheet['L1'] = 'prtReferenceNo'
sheet['M1'] = 'slaMaxDateTimeExternal'
sheet['N1'] = 'slaMaxDateTimeInternal'
sheet['O1'] = 'tglFoto'
sheet['P1'] = 'tujuanKota'
sheet['Q1'] = 'deliveryDate'
sheet['R1'] = 'statusSLA'



es = Elasticsearch(hosts="http://sat:mrrLn6gX@localhost:9200")
print(es)

awbNumber = ["000835633844",
             "001377000140",
             "000473380599",
             "000235099890",
             "000988015675"]

for awb in awbNumber:
    res = es.search(index='hghlpb', body ={
            "size": 100,
            "_source": {
                "excludes": []
            },
            "stored_fields": [
                "*"
            ],
            "script_fields": {
                "deliveryDate": {
                    "script": {
                        "source": "if (doc['status'].size() <= 0 || doc['lastValidTrackingDateTime'].size() <= 0) {\n    return null;\n}\n\ndef status = doc['status'].value;\ndef lastStatusDateTime = doc['lastValidTrackingDateTime'].value.withZoneSameInstant(ZoneId.of('Z'));\n\nif (status == 1) {\n    return lastStatusDateTime;\n} else {\n    return null;   \n}\n",
                        "lang": "painless"
                    }
                },
                "statusSLA": {
                    "script": {
                        "source": "if (doc['SlaMaxDateTimeInternal'].size() <= 0 || doc['lastValidTrackingDateTime'].size() <= 0) {\n    return \"IN PROGRESS\";\n}\n\nInstant nowInstant = Instant.ofEpochMilli(new Date().getTime());\ndef now = ZonedDateTime.ofInstant(nowInstant, ZoneId.of('Z')).plusMinutes(420);\ndef slaDate = doc['SlaMaxDateTimeInternal'].value.withZoneSameInstant(ZoneId.of('Z'));\ndef lastTrackingDate = doc['lastValidTrackingDateTime'].value.withZoneSameInstant(ZoneId.of('Z'));\ndef status = doc['status'].value;\n\ndef nowDiff = ChronoUnit.DAYS.between(slaDate, now);\ndef lastStatusDiff = ChronoUnit.DAYS.between(slaDate, lastTrackingDate);\n\nif((status == 0 && nowDiff > 0) || (status == 1 && lastStatusDiff > 0)) {\n    return \"OVER SLA\";\n}\nelse{\n    return \"MEET SLA\";\n}",
                        "lang": "painless"
                    }
                }
            },
            "query": {
                "bool": {
                    "must": [],
                    "filter": [{
                        "terms": {
                            "awbNumber": [
    awb
                            ]
                        }
                    }
                    ],
                    "should": [],
                    "must_not": []
                }
            }

    })


#print(res['hits']['hits'])
#print()
#print(res['hits'])

    content = ([])

    for hit in res['hits']['hits']:
        lastValidTrackingDateTime = dateutil.parser.parse(hit['_source']['lastValidTrackingDateTime'])



        content = (
            [hit['_source']['asalKota'], hit['_source']['awbNumber'], lastValidTrackingDateTime.strftime('%m/%d/%Y %H:%M:%S %p'),
             hit['_source']['lastValidTrackingSiteName'],
             hit['_source']['lastValidTrackingType'], hit['_source']['layanan'], hit['_source']['manifestDateTime'],
             hit['_source']['manifestTrackingSiteCity'], hit['_source']['manifestTrackingSiteName'], hit['_source']['noKonfirmasi'],
             hit['_source']['perwakilan'], hit['_source']['prtReferenceNo'], hit['_source']['SlaMaxDateTimeExternal'],
             hit['_source']['SlaMaxDateTimeInternal'], hit['_source']['tglFoto'], hit['_source']['tujuanKota'], ''.join(hit['fields']['deliveryDate']), ''.join(hit['fields']['statusSLA'])
             ],
        )
        print(hit['_source']['asalKota'])
        print(hit['_source']['awbNumber'])
        print(hit['_source']['lastValidTrackingDateTime'])
        print(hit['_source']['lastValidTrackingSiteName'])
        print(hit['_source']['lastValidTrackingType'])
        print(hit['_source']['layanan'])
        print(hit['_source']['manifestDateTime'])
        print(hit['_source']['manifestTrackingSiteCity'])
        print(hit['_source']['manifestTrackingSiteName'])
        print(hit['_source']['noKonfirmasi'])
        print(hit['_source']['perwakilan'])
        print(hit['_source']['prtReferenceNo'])
        print(hit['_source']['SlaMaxDateTimeExternal'])
        print(hit['_source']['SlaMaxDateTimeInternal'])
        print(hit['_source']['tglFoto'])
        print(hit['_source']['tujuanKota'])
    

    for i in content:
        sheet.append(i)
        print(i)


wb.save('test_Excel2.xlsx')
